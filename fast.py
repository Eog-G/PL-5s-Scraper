from re import L
import requests
import random
import shutil
import pandas as pd
from os.path import exists
import openpyxl
import time

class Player():

    def __init__(self, forename, surname, number, country, position) -> None:
        
        self.forename = forename
        self.surname = surname
        self.number = number
        self.country = country
        self.position = position
        
def get_pl_players():

    pl_players = []

    path = "players.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    maxRows = sheet.max_row

    for x in range(1,maxRows):
        prePlayer = []

        for y in range(1,6):
            prePlayer.append(sheet.cell(row=x, column=y).value)
        
        new_player = Player(prePlayer[0], prePlayer[1], prePlayer[3], prePlayer[2], prePlayer[4])

        pl_players.append(new_player)

    return pl_players

def sort_players(pl_players):

    GK = []
    DF = []
    MF = []
    FW = []

    for x in pl_players:
        
        match x.position:
            
            case "GK":
                GK.append(x)
            case "DF":
                DF.append(x)
            case "MF":
                MF.append(x)
            case "FW":
                FW.append(x)
    
    return GK,DF,MF,FW

def generate_5s_team(GK,DF,MF,FW):

    gk = GK[random.randint(0,len(GK)-1)]
    df = DF[random.randint(0,len(DF)-1)]
    mf1 = MF[random.randint(0,len(MF)-1)]
    mf2 = MF[random.randint(0,len(MF)-1)]
    fw = FW[random.randint(0,len(FW)-1)]

    return gk, df, mf1, mf2, fw

def print_5s_team(gk, df, mf1, mf2, fw):
    print_centre(F"{fw.forename[0]}{fw.surname}")
    print("")
    print_centre(F"{mf1.forename[0]}{mf1.surname}  {mf2.forename[0]}{mf2.surname}")
    print("")
    print_centre(F"{df.forename[0]}{df.surname}")
    print("")
    print_centre(F"{gk.forename[0]}{gk.surname}")


def print_centre(s):
    print(s.center(shutil.get_terminal_size().columns))

def save_data(pl_players):  
    df = pd.DataFrame([(x.forename, x.surname, x.country, x.number, x.position) for x in pl_players], columns=['Forename', 'Surname', 'Country', 'Number', 'Position'])
    df.to_excel('players.xlsx', sheet_name='players', index=False)

def main():
    start = time.time()
    pl_players = get_pl_players()

    GK, DF, MF, FW = sort_players(pl_players)
    gk_1, df_1, mf1_1, mf2_1, fw_1 = generate_5s_team(GK, DF, MF, FW)
    gk_2, df_2, mf1_2, mf2_2, fw_2 = generate_5s_team(GK, DF, MF, FW)
    
    uInput = 1

    while uInput in range(1,3):
        
        print_5s_team(gk_1, df_1, mf1_1, mf2_1, fw_1)
        print("\n")
        print_centre("----- VS -----")
        print("\n")
        print_5s_team(gk_2, df_2, mf1_2, mf2_2, fw_2)
    
        uInput = int(input("1 Wins? : 2 Wins?: "))

        if(uInput == 1):
            gk_2, df_2, mf1_2, mf2_2, fw_2 = generate_5s_team(GK, DF, MF, FW)
        if(uInput == 2):
            gk_1, df_1, mf1_1, mf2_1, fw_1 = generate_5s_team(GK, DF, MF, FW)
        

    end = time.time()
    
    print(end - start)

    if(not exists("players.xlsx")):
        save_data(pl_players)
        
main()


